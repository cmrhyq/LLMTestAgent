"""
Excel导出模块

将测试用例导出为Excel文件，支持：
- 用例清单Sheet
- 依赖关系Sheet
- 格式美化（JSON换行、必填项标红、表头冻结）
"""

import json
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, 
    Alignment, 
    PatternFill, 
    Border, 
    Side,
    NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import TestCase, TestResult, Priority
from src.core.config import get_config, AppConfig
from src.core.logging import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """
    Excel导出器
    
    将测试用例和测试结果导出为Excel文件。
    
    Attributes:
        config: 应用配置
        workbook: Excel工作簿
    """
    
    # 列定义
    CASE_COLUMNS = [
        ("用例ID", "case_id", 20),
        ("用例名称", "case_name", 30),
        ("优先级", "priority", 10),
        ("API地址", "api_url", 40),
        ("请求方法", "method", 10),
        ("请求头", "headers", 40),
        ("请求体", "body", 50),
        ("断言规则", "assert_rules", 40),
        ("依赖接口ID", "dependencies", 20),
        ("预期结果", "expected_result", 15),
        ("备注", "remark", 30),
    ]
    
    DEPENDENCY_COLUMNS = [
        ("用例ID", "case_id", 20),
        ("依赖接口ID", "dep_id", 20),
        ("依赖参数路径", "source_path", 30),
        ("目标参数位置", "target_param", 30),
    ]
    
    RESULT_COLUMNS = [
        ("用例ID", "case_id", 20),
        ("用例名称", "case_name", 30),
        ("状态", "status", 10),
        ("请求方法", "request_method", 12),
        ("请求地址", "request_url", 45),
        ("请求头", "request_headers", 40),
        ("请求数据", "request_body", 45),
        ("响应状态码", "response_status_code", 15),
        ("响应头", "response_headers", 40),
        ("响应数据", "response_body", 50),
        ("响应时间(ms)", "response_time", 15),
        ("断言结果", "assert_results", 40),
        ("错误信息", "error_message", 50),
        ("开始时间", "started_at", 20),
        ("结束时间", "finished_at", 20),
    ]
    
    def __init__(self, config: Optional[AppConfig] = None):
        """
        初始化Excel导出器
        
        Args:
            config: 应用配置
        """
        self.config = config or get_config()
        self.workbook: Optional[Workbook] = None
        self._init_styles()
    
    def _init_styles(self) -> None:
        """初始化样式"""
        # 表头样式
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 数据单元格样式
        self.data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # 必填项样式（红色背景）
        self.required_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 优先级样式
        self.p0_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        self.p1_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
        self.p2_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        
        # 状态样式
        self.passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.skipped_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # 边框样式
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    
    def export_test_cases(
        self, 
        cases: List[TestCase], 
        output_path: Optional[str] = None
    ) -> str:
        """
        导出测试用例到Excel
        
        Args:
            cases: 测试用例列表
            output_path: 输出路径，如果为None则自动生成
            
        Returns:
            str: Excel文件路径
        """
        self.workbook = Workbook()
        
        # 创建用例清单Sheet
        self._create_case_sheet(cases)
        
        # 创建依赖关系Sheet
        self._create_dependency_sheet(cases)
        
        # 生成输出路径
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = Path(self.config.output.test_cases_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = str(output_dir / f"test_cases_{timestamp}.xlsx")
        
        # 保存文件
        self.workbook.save(output_path)
        logger.info(f"测试用例已导出到: {output_path}")
        
        return output_path
    
    def export_test_results(
        self,
        results: List[TestResult],
        output_path: Optional[str] = None
    ) -> str:
        """
        导出测试结果到Excel
        
        Args:
            results: 测试结果列表
            output_path: 输出路径
            
        Returns:
            str: Excel文件路径
        """
        self.workbook = Workbook()
        
        # 创建结果Sheet
        self._create_result_sheet(results)
        
        # 生成输出路径
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = Path(self.config.output.reports_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = str(output_dir / f"test_results_{timestamp}.xlsx")
        
        # 保存文件
        self.workbook.save(output_path)
        logger.info(f"测试结果已导出到: {output_path}")
        
        return output_path
    
    def _create_case_sheet(self, cases: List[TestCase]) -> None:
        """
        创建用例清单Sheet
        
        Args:
            cases: 测试用例列表
        """
        ws = self.workbook.active
        ws.title = "用例清单"
        
        # 写入表头
        for col_idx, (header, _, width) in enumerate(self.CASE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 写入数据
        for row_idx, case in enumerate(cases, 2):
            self._write_case_row(ws, row_idx, case)
        
        # 冻结表头
        if self.config.output.excel.freeze_header:
            ws.freeze_panes = "A2"
        
        # 添加筛选
        ws.auto_filter.ref = ws.dimensions
    
    def _write_case_row(self, ws: Worksheet, row_idx: int, case: TestCase) -> None:
        """
        写入用例行
        
        Args:
            ws: 工作表
            row_idx: 行索引
            case: 测试用例
        """
        row_data = [
            case.case_id,
            case.case_name,
            case.priority.value,
            case.api_url,
            case.method.value,
            self._format_json(case.headers),
            self._format_json(case.body) if case.body else "",
            "; ".join(case.assert_rules),
            ", ".join(case.dependencies.keys()),
            case.expected_result,
            case.remark,
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = self.data_alignment
            cell.border = self.thin_border
            
            # 优先级着色
            if col_idx == 3:  # 优先级列
                if case.priority == Priority.P0:
                    cell.fill = self.p0_fill
                elif case.priority == Priority.P1:
                    cell.fill = self.p1_fill
                elif case.priority == Priority.P2:
                    cell.fill = self.p2_fill
    
    def _create_dependency_sheet(self, cases: List[TestCase]) -> None:
        """
        创建依赖关系Sheet
        
        Args:
            cases: 测试用例列表
        """
        ws = self.workbook.create_sheet("依赖关系")
        
        # 写入表头
        for col_idx, (header, _, width) in enumerate(self.DEPENDENCY_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 写入数据
        row_idx = 2
        for case in cases:
            for dep_id, dep_info in case.dependencies.items():
                row_data = [
                    case.case_id,
                    dep_id,
                    dep_info.get("source_path", ""),
                    dep_info.get("target_param", ""),
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = self.data_alignment
                    cell.border = self.thin_border
                row_idx += 1
        
        # 冻结表头
        if self.config.output.excel.freeze_header:
            ws.freeze_panes = "A2"
    
    def _create_result_sheet(self, results: List[TestResult]) -> None:
        """
        创建测试结果Sheet
        
        Args:
            results: 测试结果列表
        """
        ws = self.workbook.active
        ws.title = "测试结果"
        
        # 写入表头
        for col_idx, (header, _, width) in enumerate(self.RESULT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 写入数据
        for row_idx, result in enumerate(results, 2):
            self._write_result_row(ws, row_idx, result)
        
        # 冻结表头
        if self.config.output.excel.freeze_header:
            ws.freeze_panes = "A2"
        
        # 添加筛选
        ws.auto_filter.ref = ws.dimensions
    
    def _write_result_row(self, ws: Worksheet, row_idx: int, result: TestResult) -> None:
        """
        写入结果行
        
        Args:
            ws: 工作表
            row_idx: 行索引
            result: 测试结果
        """
        row_data = [
            result.case_id,
            result.case_name,
            result.status.value,
            result.request_method,
            result.request_url,
            self._format_json(result.request_headers) if result.request_headers else "",
            self._format_json(result.request_body) if result.request_body else "",
            result.response_status_code or "",
            self._format_json(result.response_headers) if result.response_headers else "",
            self._format_json(result.response_body) if result.response_body is not None else "",
            round(result.response_time, 2) if result.response_time else "",
            self._format_json(result.assert_results) if result.assert_results else "",
            result.error_message,
            result.started_at.strftime("%Y-%m-%d %H:%M:%S") if result.started_at else "",
            result.finished_at.strftime("%Y-%m-%d %H:%M:%S") if result.finished_at else "",
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = self.data_alignment
            cell.border = self.thin_border
            
            # 状态着色
            if col_idx == 3:  # 状态列
                if result.status.value == "passed":
                    cell.fill = self.passed_fill
                elif result.status.value in ("failed", "error"):
                    cell.fill = self.failed_fill
                elif result.status.value == "skipped":
                    cell.fill = self.skipped_fill
    
    def _format_json(self, data: Any) -> str:
        """
        格式化JSON数据
        
        Args:
            data: 数据
            
        Returns:
            str: 格式化后的JSON字符串
        """
        if data is None:
            return ""
        if isinstance(data, str):
            return data
        try:
            return json.dumps(data, ensure_ascii=False, indent=2)
        except Exception:
            return str(data)


def export_test_cases(
    cases: List[TestCase],
    output_path: Optional[str] = None,
    config: Optional[AppConfig] = None
) -> str:
    """
    导出测试用例的便捷函数
    
    Args:
        cases: 测试用例列表
        output_path: 输出路径
        config: 应用配置
        
    Returns:
        str: Excel文件路径
    """
    exporter = ExcelExporter(config)
    return exporter.export_test_cases(cases, output_path)


def export_test_results(
    results: List[TestResult],
    output_path: Optional[str] = None,
    config: Optional[AppConfig] = None
) -> str:
    """
    导出测试结果的便捷函数
    
    Args:
        results: 测试结果列表
        output_path: 输出路径
        config: 应用配置
        
    Returns:
        str: Excel文件路径
    """
    exporter = ExcelExporter(config)
    return exporter.export_test_results(results, output_path)
